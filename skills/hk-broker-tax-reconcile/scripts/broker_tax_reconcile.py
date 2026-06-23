#!/usr/bin/env python3
"""Offline HK broker tax reconciliation helper.

This script intentionally performs local-only parsing. It never uploads source
files and it does not embed any user identity data.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, getcontext
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required to parse XLSX broker files") from exc

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover
    PdfReader = None


getcontext().prec = 28
CENT = Decimal("0.01")


def d(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value).replace(",", ""))


def q2(value: Decimal) -> Decimal:
    return value.quantize(CENT, rounding=ROUND_HALF_UP)


def norm_ccy(ccy: str) -> str:
    return "CNH" if ccy in ("CNY", "CNH") else ccy


def parse_dt(value) -> datetime:
    text = str(value)
    if "-" in text:
        return datetime.strptime(text[:10], "%Y-%m-%d")
    return datetime.strptime(text[:8], "%Y%m%d")


def sheet_rows(wb, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        return []
    values = list(wb[sheet_name].iter_rows(values_only=True))
    if not values:
        return []
    header = [str(cell) for cell in values[0]]
    return [dict(zip(header, row)) for row in values[1:] if any(cell not in (None, "") for cell in row)]


@dataclass
class Lot:
    broker: str
    account: str
    code: str
    market: str
    currency: str
    qty: Decimal
    total_cost: Decimal
    fee: Decimal = Decimal("0")
    buy_date: str = ""
    note: str = ""

    @property
    def cost_with_fee(self) -> Decimal:
        return self.total_cost + self.fee


@dataclass
class Position:
    qty: Decimal = Decimal("0")
    cost: Decimal = Decimal("0")
    sources: defaultdict[str, Decimal] = field(default_factory=lambda: defaultdict(Decimal))


def load_manual_lots(path: Path | None) -> dict[tuple[str, str, str, str], list[Lot]]:
    lots: dict[tuple[str, str, str, str], list[Lot]] = defaultdict(list)
    if not path:
        return lots
    with path.open(encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            status = (row.get("status") or "confirmed").strip().lower()
            if status not in ("", "confirmed", "ok"):
                continue
            broker = (row.get("broker") or "").strip().lower()
            account = (row.get("account") or "").strip()
            code = (row.get("code") or "").strip()
            currency = norm_ccy((row.get("currency") or "").strip())
            market = (row.get("market") or "").strip()
            qty = d(row.get("qty"))
            total_cost = d(row.get("total_cost"))
            if total_cost == 0 and row.get("unit_cost"):
                total_cost = qty * d(row.get("unit_cost"))
            lots[(broker, account, code, currency)].append(
                Lot(
                    broker=broker,
                    account=account,
                    code=code,
                    market=market,
                    currency=currency,
                    qty=qty,
                    total_cost=total_cost,
                    fee=d(row.get("fee")),
                    buy_date=(row.get("buy_date") or "").strip(),
                    note=(row.get("source_note") or row.get("note") or "").strip(),
                )
            )
    return lots


def manual_cost_for(
    lots: dict[tuple[str, str, str, str], list[Lot]],
    broker: str,
    account: str,
    code: str,
    currency: str,
) -> tuple[Decimal, Decimal]:
    rows = lots.get((broker, account, code, norm_ccy(currency)), [])
    return sum((lot.qty for lot in rows), Decimal("0")), sum((lot.cost_with_fee for lot in rows), Decimal("0"))


def replay_events(events: list[tuple]) -> tuple[list[dict], list[str]]:
    events.sort(key=lambda item: item[0])
    positions: defaultdict[tuple[str, str], Position] = defaultdict(Position)
    realized: list[dict] = []
    warnings: list[str] = []

    for event_dt, kind, code, ccy, qty, amount, source in events:
        key = (code, ccy)
        pos = positions[key]
        if kind == "open":
            pos.qty += qty
            pos.cost += amount
            pos.sources[source] += qty
            continue

        if pos.qty < qty:
            warnings.append(
                f"{event_dt.date()} {code}/{ccy}: sell quantity {qty} exceeds available quantity {pos.qty}"
            )
            continue

        avg_cost = pos.cost / pos.qty if pos.qty else Decimal("0")
        cost = avg_cost * qty
        pnl = amount - cost
        realized.append(
            {
                "date": event_dt.date().isoformat(),
                "code": code,
                "currency": ccy,
                "proceeds_net": amount,
                "cost": cost,
                "realized_pnl": pnl,
                "sources": ";".join(sorted(pos.sources.keys())),
            }
        )

        ratio = qty / pos.qty
        pos.qty -= qty
        pos.cost -= cost
        for src in list(pos.sources):
            pos.sources[src] -= pos.sources[src] * ratio
            if abs(pos.sources[src]) < Decimal("0.00000001"):
                del pos.sources[src]

    return realized, warnings


def parse_futu_xlsx(path: Path, manual_lots: dict) -> tuple[list[dict], list[dict], list[dict], list[str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    holdings = sheet_rows(wb, "证券-持仓总览")
    trades = sheet_rows(wb, "证券-交易流水")
    assets = sheet_rows(wb, "证券-资产进出")
    cashflows = sheet_rows(wb, "证券-资金进出")
    broker = "futu"

    ipo_cash: defaultdict[str, Decimal] = defaultdict(Decimal)
    for row in cashflows:
        if row.get("类型") == "港股IPO公开发售":
            match = re.search(r"#(\d{5})", str(row.get("备注") or ""))
            if match:
                ipo_cash[match.group(1).lstrip("0")] += d(row.get("变动金额"))

    events: list[tuple] = []
    missing_costs: list[dict] = []
    for row in holdings:
        if row.get("时期类型") != "期初":
            continue
        qty = d(row.get("数量/面值"))
        if qty == 0:
            continue
        account = str(row.get("账户号码") or "")
        code = str(row.get("代码名称") or "")
        ccy = norm_ccy(str(row.get("币种") or ""))
        manual_qty, manual_cost = manual_cost_for(manual_lots, broker, account, code, ccy)
        source = "manual_opening_cost"
        if manual_qty == qty and manual_cost > 0:
            cost = manual_cost
        else:
            source = "opening_placeholder_not_final"
            cost = qty * d(row.get("价格"))
            missing_costs.append(
                {
                    "broker": broker,
                    "account": account,
                    "code": code,
                    "market": row.get("交易所/市场"),
                    "currency": ccy,
                    "opening_date": row.get("日期"),
                    "opening_qty": str(qty),
                    "placeholder_price": str(row.get("价格")),
                    "required_fields": "buy_date, qty, unit_cost or total_cost, fee",
                }
            )
        events.append((datetime(1900, 1, 1), "open", code, ccy, qty, cost, source))

    for row in assets:
        if row.get("类型") == "港股IPO公开发售" and row.get("方向") == "In":
            code = str(row.get("代码名称") or "")
            qty = d(row.get("数量"))
            net_cash = ipo_cash.get(code, Decimal("0"))
            if qty > 0 and net_cash < 0:
                events.append(
                    (parse_dt(row.get("日期")), "open", code, norm_ccy(str(row.get("币种") or "")), qty, -net_cash, "ipo_cash_net")
                )

    for row in trades:
        qty = d(row.get("数量/面值"))
        code = str(row.get("代码名称") or "")
        ccy = norm_ccy(str(row.get("币种") or ""))
        trade_dt = parse_dt(row.get("成交时间"))
        if qty > 0:
            events.append((trade_dt, "open", code, ccy, qty, -d(row.get("变动金额")), "trade_buy"))
        elif qty < 0:
            events.append((trade_dt, "close", code, ccy, -qty, d(row.get("变动金额")), "trade_sell_net"))

    realized, warnings = replay_events(events)
    for row in realized:
        row["broker"] = broker

    dividends = parse_futu_dividends(cashflows)
    return realized, dividends, missing_costs, warnings


def parse_futu_dividends(cashflows: Iterable[dict]) -> list[dict]:
    out: list[dict] = []
    for row in cashflows:
        if row.get("类型") != "公司行动":
            continue
        amount = d(row.get("变动金额"))
        ccy = norm_ccy(str(row.get("币种") or ""))
        note = str(row.get("备注") or "")
        upper = note.upper()
        gross = withholding = fee = Decimal("0")
        kind = None
        if "WITHHOLDING" in upper or "- TAX" in upper:
            kind = "withholding"
            withholding = -amount if amount < 0 else amount
        elif "HANDLING CHARGE" in upper or "SCRIP CHARGE" in upper:
            kind = "fee"
            fee = -amount if amount < 0 else amount
        elif any(marker in upper for marker in ("DIVIDENDS", "F/D", "S/D", "I/D")):
            kind = "dividend"
            if "(-10%)" in upper or "10% TAX INCLUDED" in upper:
                gross = amount / Decimal("0.9")
                withholding = gross - amount
            else:
                gross = amount
        if kind:
            out.append(
                {
                    "broker": "futu",
                    "date": row.get("日期"),
                    "currency": ccy,
                    "type": kind,
                    "cash_amount": amount,
                    "gross_dividend": gross,
                    "withholding_tax": withholding,
                    "fee": fee,
                    "note": note,
                }
            )
    return out


def parse_hafoo_xlsx(path: Path) -> tuple[list[dict], list[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    pl_rows = sheet_rows(wb, "个股盈亏")
    trade_rows = sheet_rows(wb, "交易明细")
    dividend_rows = sheet_rows(wb, "分红派息")
    pl_by_code = {str(row.get("代码")): d(row.get("累计实现盈亏")) for row in pl_rows}

    realized: list[dict] = []
    for row in trade_rows:
        code = str(row.get("代码"))
        proceeds = d(row.get("清算金额"))
        pnl = pl_by_code.get(code, Decimal("0"))
        realized.append(
            {
                "broker": "hafoo",
                "date": str(row.get("日期")),
                "code": code,
                "currency": norm_ccy(str(row.get("币种") or "")),
                "proceeds_net": proceeds,
                "cost": proceeds - pnl,
                "realized_pnl": pnl,
                "sources": "broker_reported_realized_pnl",
            }
        )

    dividends: list[dict] = []
    for row in dividend_rows:
        dividends.append(
            {
                "broker": "hafoo",
                "date": row.get("日期"),
                "currency": norm_ccy(str(row.get("币种") or "")),
                "type": "dividend",
                "cash_amount": d(row.get("股息金额（税后）")),
                "gross_dividend": d(row.get("股息金额（税前）")),
                "withholding_tax": d(row.get("预扣税金额")),
                "fee": Decimal("0"),
                "note": row.get("备注") or "",
            }
        )
    return realized, dividends


def load_manual_dividends(path: Path | None) -> list[dict]:
    if not path:
        return []
    out: list[dict] = []
    with path.open(encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            out.append(
                {
                    "broker": row.get("broker") or "manual",
                    "date": row.get("date") or "",
                    "currency": norm_ccy(row.get("currency") or ""),
                    "type": row.get("type") or "dividend",
                    "cash_amount": d(row.get("net_cash")),
                    "gross_dividend": d(row.get("gross_dividend")),
                    "withholding_tax": d(row.get("withholding_tax")),
                    "fee": d(row.get("fee")),
                    "note": row.get("note") or "",
                }
            )
    return out


def inspect_pdf(path: Path) -> dict:
    if PdfReader is None:
        return {"file": path.name, "error": "pypdf not installed"}
    reader = PdfReader(str(path))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    return {
        "file": path.name,
        "pages": len(reader.pages),
        "contains_activity_statement": "活动报表" in text,
        "contains_tax_form": "税表" in text or "Tax" in text,
        "contains_calendar_2025": "2025.01.01 - 2025.12.31" in text or "2025-01-01" in text,
    }


def parse_tiger_activity_pdf(path: Path) -> tuple[list[dict], dict]:
    """Parse common Tiger activity-statement dividend rows from a PDF.

    Tiger PDF extraction is layout-sensitive, so this parser captures the
    dividend patterns that usually survive text extraction and returns an
    inspection report alongside parsed rows. If no rows are found, callers can
    fall back to a manual dividend summary CSV.
    """
    inspection = inspect_pdf(path)
    if PdfReader is None:
        return [], inspection

    reader = PdfReader(str(path))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    rows: list[dict] = []

    dividend_pattern = re.compile(
        r"(?P<date>\d{4}-\d{2}-\d{2})\s+股票\s+(?P<name>.*?)\n"
        r"\((?P<code>[^)]+)\).*?"
        r"派发\s+(?P<gross>[\d,]+(?:\.\d+)?)\s+0\s+公司行动手续费:\s+(?P<fee>[\d,]+(?:\.\d+)?)\n"
        r"分红税:\s+(?P<tax>[\d,]+(?:\.\d+)?)\s+(?P<net>[\d,]+(?:\.\d+)?)\s+(?P<currency>[A-Z]{3})",
        re.S,
    )
    for match in dividend_pattern.finditer(text):
        rows.append(
            {
                "broker": "tiger",
                "date": match.group("date"),
                "currency": norm_ccy(match.group("currency")),
                "type": "dividend",
                "cash_amount": d(match.group("net")),
                "gross_dividend": d(match.group("gross")),
                "withholding_tax": d(match.group("tax")),
                "fee": d(match.group("fee")),
                "note": f"{match.group('code')} {match.group('name').strip()}",
            }
        )

    adr_pattern = re.compile(
        r"(?P<date>\d{4}-\d{2}-\d{2})\s+股票\s+(?P<name>.*?)\n"
        r"\((?P<code>[^)]+)\).*?"
        r"ADR费用:\s+(?P<fee>[\d,]+(?:\.\d+)?)\s+(?P<net>-?[\d,]+(?:\.\d+)?)\s+(?P<currency>[A-Z]{3})",
        re.S,
    )
    for match in adr_pattern.finditer(text):
        rows.append(
            {
                "broker": "tiger",
                "date": match.group("date"),
                "currency": norm_ccy(match.group("currency")),
                "type": "adr_fee",
                "cash_amount": d(match.group("net")),
                "gross_dividend": Decimal("0"),
                "withholding_tax": Decimal("0"),
                "fee": d(match.group("fee")),
                "note": f"{match.group('code')} {match.group('name').strip()}",
            }
        )

    inspection["parsed_dividend_rows"] = len(rows)
    return rows, inspection


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def to_hkd(amount: Decimal, currency: str, hkd_cny: Decimal, usd_hkd: Decimal) -> Decimal:
    currency = norm_ccy(currency)
    if currency == "HKD":
        return amount
    if currency in ("CNH", "CNY"):
        return amount / hkd_cny
    if currency == "USD":
        return amount * usd_hkd
    return amount


def build_tax_summary(
    realized: list[dict],
    dividends: list[dict],
    hkd_cny: Decimal,
    usd_hkd: Decimal,
) -> dict:
    capital_gain_hkd = sum((to_hkd(d(row["realized_pnl"]), row["currency"], hkd_cny, usd_hkd) for row in realized), Decimal("0"))
    dividend_gross_hkd = sum((to_hkd(d(row["gross_dividend"]), row["currency"], hkd_cny, usd_hkd) for row in dividends), Decimal("0"))
    dividend_withheld_hkd = sum((to_hkd(d(row["withholding_tax"]), row["currency"], hkd_cny, usd_hkd) for row in dividends), Decimal("0"))
    capital_tax_hkd = capital_gain_hkd * Decimal("0.20")
    dividend_tax_hkd = dividend_gross_hkd * Decimal("0.20")
    dividend_makeup_hkd = max(Decimal("0"), dividend_tax_hkd - dividend_withheld_hkd)
    return {
        "rates": {"HKD_CNY": str(hkd_cny), "USD_HKD": str(usd_hkd)},
        "capital_gain_hkd": str(q2(capital_gain_hkd)),
        "capital_gain_cny": str(q2(capital_gain_hkd * hkd_cny)),
        "capital_tax_20pct_cny": str(q2(capital_tax_hkd * hkd_cny)),
        "dividend_gross_hkd": str(q2(dividend_gross_hkd)),
        "dividend_gross_cny": str(q2(dividend_gross_hkd * hkd_cny)),
        "dividend_withholding_cny": str(q2(dividend_withheld_hkd * hkd_cny)),
        "dividend_makeup_tax_cny": str(q2(dividend_makeup_hkd * hkd_cny)),
        "filing_fields": {
            "property_transfer_taxable_income_cny": str(q2(capital_gain_hkd * hkd_cny)),
            "interest_dividend_bonus_taxable_income_cny": str(q2(dividend_gross_hkd * hkd_cny)),
            "current_year_foreign_tax_paid_cny": str(q2(dividend_withheld_hkd * hkd_cny)),
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Reconcile local broker files into tax filing CSVs.")
    parser.add_argument("--futu-xlsx", type=Path)
    parser.add_argument("--hafoo-xlsx", type=Path)
    parser.add_argument("--tiger-activity-pdf", type=Path, action="append", default=[])
    parser.add_argument("--manual-lots", type=Path, help="CSV with manual opening/transfer lots")
    parser.add_argument("--manual-dividends", type=Path, help="CSV with manually summarized dividends, useful for Tiger PDFs")
    parser.add_argument("--hkd-cny", type=Decimal, required=True, help="Same-date HKD to CNY rate")
    parser.add_argument("--usd-hkd", type=Decimal, required=True, help="Same-date USD to HKD rate")
    parser.add_argument("--out-dir", type=Path, default=Path("outputs/broker_tax_reconcile_run"))
    args = parser.parse_args()

    args.out_dir.mkdir(parents=True, exist_ok=True)
    manual_lots = load_manual_lots(args.manual_lots)
    realized: list[dict] = []
    dividends: list[dict] = []
    missing_costs: list[dict] = []
    warnings: list[str] = []
    inspections: list[dict] = []

    if args.futu_xlsx:
        futu_realized, futu_dividends, futu_missing, futu_warnings = parse_futu_xlsx(args.futu_xlsx, manual_lots)
        realized.extend(futu_realized)
        dividends.extend(futu_dividends)
        missing_costs.extend(futu_missing)
        warnings.extend(futu_warnings)

    if args.hafoo_xlsx:
        hafoo_realized, hafoo_dividends = parse_hafoo_xlsx(args.hafoo_xlsx)
        realized.extend(hafoo_realized)
        dividends.extend(hafoo_dividends)

    for pdf_path in args.tiger_activity_pdf:
        tiger_dividends, inspection = parse_tiger_activity_pdf(pdf_path)
        dividends.extend(tiger_dividends)
        inspections.append(inspection)

    dividends.extend(load_manual_dividends(args.manual_dividends))

    write_csv(
        args.out_dir / "realized_capital_gains.csv",
        realized,
        ["broker", "date", "code", "currency", "proceeds_net", "cost", "realized_pnl", "sources"],
    )
    write_csv(
        args.out_dir / "dividend_withholding_items.csv",
        dividends,
        ["broker", "date", "currency", "type", "cash_amount", "gross_dividend", "withholding_tax", "fee", "note"],
    )
    write_csv(
        args.out_dir / "manual_lots_needed.csv",
        missing_costs,
        ["broker", "account", "code", "market", "currency", "opening_date", "opening_qty", "placeholder_price", "required_fields"],
    )
    summary = build_tax_summary(realized, dividends, args.hkd_cny, args.usd_hkd)
    summary["final_ready"] = not missing_costs and not warnings
    summary["warnings"] = warnings
    summary["pdf_inspections"] = inspections
    (args.out_dir / "tax_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
